import asyncio
import logging
import re
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from openpyxl import load_workbook

# ===== НАСТРОЙКИ =====
BOT_TOKEN = "8464995539:AAEtJDETZbsYYR2W5fnIlZeElxplXLO2zAQ"          # замените на токен от @BotFather
EXCEL_FILE = "Цены 1.xlsx"

# Стоимость доставки и забора
INSIDE_MKAD = 3000      # внутри МКАД (руб)
OUTSIDE_MKAD = 5000     # за МКАД (руб)

# ===== НАСТРОЙКА ЛОГИРОВАНИЯ =====
logging.basicConfig(level=logging.INFO)

# ===== ИНИЦИАЛИЗАЦИЯ БОТА =====
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

# ===== ГЛОБАЛЬНЫЙ СЛОВАРЬ ДАННЫХ =====
cars_by_id = {}          # словарь {идентификатор: данные автомобиля}

# ===== ФУНКЦИЯ ФОРМАТИРОВАНИЯ ЧИСЕЛ =====
def format_number(value):
    """Преобразует число в строку с пробелами между разрядами.
       Если value не число или None, возвращает исходное значение или "уточняйте"."""
    if value is None:
        return "уточняйте"
    try:
        # Пробуем преобразовать в число (int или float) и затем в int
        num = int(float(value))
        return f"{num:,}".replace(',', ' ')
    except (ValueError, TypeError):
        # Если не число, возвращаем как есть (например, строку)
        return str(value)

# ===== ФУНКЦИЯ ЗАГРУЗКИ ДАННЫХ ИЗ EXCEL =====
def load_cars_from_excel():
    """Загружает данные из листа 'Лист1' в словарь cars_by_id."""
    wb = load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb["Лист1"]

    # Ожидаемые колонки (нумерация с 0):
    # A: Модель
    # B: Номер (идентификатор)
    # C: ссылка (URL или комментарий)
    # D: Год выпуска
    # E: Пробег в сутки
    # F: 1 сутки
    # G: 2-3 суток
    # H: 4-6 суток
    # I: 7-15 суток
    # J: 16-29 суток
    # K: 30 суток
    # L: Залог

    for row in sheet.iter_rows(min_row=2, values_only=True):  # пропускаем заголовок
        # Проверяем, что есть номер
        if not row[1]:
            continue

        car_id = str(row[1]).strip()          # столбец B
        model = str(row[0]).strip() if row[0] else ""
        raw_url = str(row[2]).strip() if row[2] else ""
        year = str(row[3]).strip() if row[3] else ""
        mileage = row[4]                       # пробег

        # Преобразуем пробег в число
        try:
            mileage = int(float(mileage)) if mileage else 250
        except:
            mileage = 250

        # Цены (приводим к int)
        def safe_int(val):
            try:
                return int(float(val)) if val is not None else None
            except:
                return None

        price_1 = safe_int(row[5])      # F
        price_2_3 = safe_int(row[6])    # G
        price_4_6 = safe_int(row[7])    # H
        price_7_15 = safe_int(row[8])   # I
        price_16_29 = safe_int(row[9])  # J
        price_month = safe_int(row[10]) # K
        deposit = safe_int(row[11])     # L

        # Проверяем, является ли ссылка валидной (начинается с http)
        if raw_url.startswith(('http://', 'https://')):
            url = raw_url
        else:
            url = None  # или можно оставить как текст "нет ссылки"

        # Сохраняем
        cars_by_id[car_id] = {
            "model": model,
            "year": year,
            "url": url,
            "mileage": mileage,
            "price_1": price_1,
            "price_2_3": price_2_3,
            "price_4_6": price_4_6,
            "price_7_15": price_7_15,
            "price_16_29": price_16_29,
            "price_month": price_month,
            "deposit": deposit,
        }

    wb.close()
    logging.info(f"Загружено {len(cars_by_id)} автомобилей.")

# ===== ФУНКЦИЯ ПОЛУЧЕНИЯ ЦЕНЫ ПО ДИАПАЗОНУ =====
def get_price_for_days(car, days):
    """Возвращает (цена_за_сутки, название_диапазона) для указанного количества дней."""
    if days <= 0:
        return None, None
    if days == 1:
        return car["price_1"], "1 сутки"
    elif 2 <= days <= 3:
        return car["price_2_3"], "2-3 суток"
    elif 4 <= days <= 6:
        return car["price_4_6"], "4-6 суток"
    elif 7 <= days <= 15:
        return car["price_7_15"], "7-15 суток"
    elif 16 <= days <= 29:
        return car["price_16_29"], "16-29 суток"
    else:  # 30 и более
        return car["price_month"], "30+ суток"

# ===== ФУНКЦИЯ ФОРМИРОВАНИЯ ОТВЕТА (ПРЕМИУМ ВАРИАНТ 2) =====
def format_car_response(car, days, range_name, price_per_day, total_price, link_text):
    total_price_f = format_number(total_price)
    price_per_day_f = format_number(price_per_day)
    deposit_f = format_number(car["deposit"])
    mileage_f = format_number(car["mileage"])
    inside_f = format_number(INSIDE_MKAD)
    outside_f = format_number(OUTSIDE_MKAD)

    text = (
        f"🚗 <b>{car['model']}</b> <i>{car['year']} г.</i>\n"
        f"🔗 {link_text}\n"   # Убрано \n\n, теперь только \n
        f"📅 <b>Срок аренды:</b> {days} дн. ({range_name})\n\n"
        f"💰 <b>Стоимость аренды:</b>\n"
        f"   ├ Общая:      <b>{total_price_f}</b> руб.\n"
        f"   ├ За сутки:   <b>{price_per_day_f}</b> руб.\n"
        f"   └ Залог:      <b>{deposit_f}</b> руб.\n"   # Убрано \n\n, теперь просто \n
        f"📏 <b>Лимит пробега:</b> {mileage_f} км/день\n\n"
        f"🚚 <b>Доставка и забор:</b>\n"
        f"   ├ Внутри МКАД: {inside_f} руб.\n"
        f"   └ За МКАД:     {outside_f} руб.\n\n"
        f"──────────\n"
        f"🌐 <a href='https://topcar-elite.ru'>topcar-elite.ru</a>"
    )
    return text

# ===== ОБРАБОТЧИКИ КОМАНД =====
@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    await message.answer(
        "👋 Привет! Я помогу рассчитать стоимость аренды автомобиля.\n"
        "Отправь мне сообщение в формате:\n"
        "<b>номер_автомобиля количество_дней</b>\n"
        "Например: <code>738 5</code>\n\n"
        "Номер автомобиля можно найти в нашем каталоге (трёхзначное число в названии).",
        parse_mode="HTML"
    )

@dp.message()
async def handle_car_request(message: types.Message):
    text = message.text.strip()
    logging.info(f"Получено сообщение: {text}")

    # Разделяем сообщение на части: последнее слово - дни, остальное - номер
    parts = text.rsplit(maxsplit=1)
    if len(parts) != 2:
        await message.answer("❌ Неверный формат. Нужно: номер_автомобиля количество_дней (например, 738 5)")
        return

    car_id_input, days_str = parts[0], parts[1]

    # Проверяем, что дни - число
    try:
        days = int(days_str)
    except ValueError:
        await message.answer("❌ Количество дней должно быть числом.")
        return

    # Нормализуем введённый номер (убираем лишние пробелы)
    car_id = car_id_input.strip()

    # Ищем в базе
    car = cars_by_id.get(car_id)
    if not car:
        await message.answer(f"❌ Автомобиль с номером {car_id} не найден.")
        return

    # Получаем цену за сутки
    price_per_day, range_name = get_price_for_days(car, days)
    if price_per_day is None:
        await message.answer("❌ Для указанного количества дней нет цены в прайсе.")
        return

    total_price = price_per_day * days

    # Формируем ссылку на карточку
    if car["url"]:
        link_text = f"<a href='{car['url']}'>Карточка авто на сайте</a>"
    else:
        link_text = "Карточка отсутствует"

    # Формируем ответ с красивым оформлением
    response = format_car_response(car, days, range_name, price_per_day, total_price, link_text)

    # Отправляем только текст
    await message.answer(response, parse_mode="HTML")

# ===== ЗАПУСК БОТА =====
async def main():
    # Удаляем вебхук (если был) и загружаем данные
    await bot.delete_webhook(drop_pending_updates=True)
    load_cars_from_excel()
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())